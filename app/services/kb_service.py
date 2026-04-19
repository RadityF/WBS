from __future__ import annotations

from dataclasses import dataclass
import re

from openpyxl import load_workbook
from qdrant_client.models import PointStruct

from app.config import settings
from app.services.ollama_client import OllamaClient
from app.services.qdrant_service import QdrantService


@dataclass
class KnowledgeRow:
    source_sheet: str
    source_group: str
    category: str
    text: str
    chunk_type: str
    law_ref: str
    sample_input: str
    scenario_hint: int
    keywords: list[str]


def _cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _extract_keywords(category: str, definition: str, sample: str) -> list[str]:
    base = set()
    stopwords = {
        "dan",
        "atau",
        "yang",
        "untuk",
        "dengan",
        "dalam",
        "pada",
        "dari",
        "oleh",
        "sebagai",
        "karyawan",
        "kantor",
    }

    category_tokens = re.split(r"[^a-zA-Z0-9]+", category.lower())
    for token in category_tokens:
        if len(token) >= 3 and token not in stopwords:
            base.add(token)

    text = f"{definition} {sample}".lower()
    all_tokens = re.findall(r"[a-zA-Z0-9-]+", text)
    for token in all_tokens:
        if len(token) >= 5 and token not in stopwords:
            base.add(token)

    hint_map = {
        "Gratifikasi": ["hadiah", "vendor", "bingkisan", "jam tangan"],
        "Suap (Bribery)": ["suap", "pelicin", "meloloskan", "tender"],
        "Korupsi / Fraud": ["mark-up", "markup", "penggelapan", "fraud"],
        "Konflik Kepentingan": ["keluarga", "titipan", "vendor", "benturan"],
        "Pelecehan (Harassment)": ["pelecehan", "komentar", "tidak senonoh", "sentuh", "paksa"],
        "Pencurian Data (Data Breach)": ["database", "nasabah", "copy", "flashdisk", "akses"],
        "Penyalahgunaan Aset": ["aset", "kendaraan", "operasional", "pribadi"],
    }
    for k, vals in hint_map.items():
        if k.lower() in category.lower():
            for item in vals:
                if item in text:
                    base.add(item)

    return sorted(base)


def load_knowledge_rows(xlsx_path: str) -> list[KnowledgeRow]:
    wb = load_workbook(xlsx_path)
    rows: list[KnowledgeRow] = []

    if "Pelanggaran" in wb.sheetnames:
        ws = wb["Pelanggaran"]
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            category = _cell(row[1])
            definition = _cell(row[2])
            law_ref = _cell(row[3])
            sample = _cell(row[4])
            if not category:
                continue
            keywords = _extract_keywords(category, definition, sample)

            rows.append(
                KnowledgeRow(
                    source_sheet="Pelanggaran",
                    source_group="violation",
                    category=category,
                    chunk_type="definition",
                    text=f"Kategori: {category}\nDefinisi Operasional: {definition}",
                    law_ref=law_ref,
                    sample_input=sample,
                    scenario_hint=1,
                    keywords=keywords,
                )
            )

            if law_ref:
                rows.append(
                    KnowledgeRow(
                        source_sheet="Pelanggaran",
                        source_group="violation",
                        category=category,
                        chunk_type="law_ref",
                        text=f"Kategori: {category}\nReferensi Hukum: {law_ref}",
                        law_ref=law_ref,
                        sample_input=sample,
                        scenario_hint=1,
                        keywords=keywords,
                    )
                )

            if sample:
                rows.append(
                    KnowledgeRow(
                        source_sheet="Pelanggaran",
                        source_group="violation",
                        category=category,
                        chunk_type="example",
                        text=f"Kategori: {category}\nContoh Narasi: {sample}",
                        law_ref=law_ref,
                        sample_input=sample,
                        scenario_hint=1,
                        keywords=keywords,
                    )
                )

    if "Out of Scope (Skenario 3)" in wb.sheetnames:
        ws = wb["Out of Scope (Skenario 3)"]
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            category = _cell(row[1])
            definition = _cell(row[2])
            sample = _cell(row[3])
            response_template = _cell(row[4])
            if not category:
                continue
            keywords = _extract_keywords(category, definition, sample)

            rows.append(
                KnowledgeRow(
                    source_sheet="OutOfScope",
                    source_group="oos",
                    category=category,
                    chunk_type="definition",
                    text=f"Jenis Keluhan: {category}\nDefinisi Filter: {definition}",
                    law_ref="",
                    sample_input=sample,
                    scenario_hint=3,
                    keywords=keywords,
                )
            )

            if sample:
                rows.append(
                    KnowledgeRow(
                        source_sheet="OutOfScope",
                        source_group="oos",
                        category=category,
                        chunk_type="example",
                        text=f"Jenis Keluhan: {category}\nContoh Narasi: {sample}",
                        law_ref="",
                        sample_input=sample,
                        scenario_hint=3,
                        keywords=keywords,
                    )
                )

            if response_template:
                rows.append(
                    KnowledgeRow(
                        source_sheet="OutOfScope",
                        source_group="oos",
                        category=category,
                        chunk_type="response_template",
                        text=f"Jenis Keluhan: {category}\nTemplate Respon: {response_template}",
                        law_ref="",
                        sample_input=sample,
                        scenario_hint=3,
                        keywords=keywords,
                    )
                )
    return rows


def reindex_knowledge_base() -> dict:
    rows = load_knowledge_rows(settings.kb_file)
    if not rows:
        return {"indexed": 0, "message": "Tidak ada data KB yang ditemukan"}

    ollama = OllamaClient()
    qdrant = QdrantService()

    first_vector = ollama.embed(rows[0].text)
    # Force recreate to avoid stale/wrong-dimension collections on remote Qdrant.
    qdrant.recreate_collection(len(first_vector), settings.qdrant_collection_violation)
    qdrant.recreate_collection(len(first_vector), settings.qdrant_collection_oos)

    violation_points: list[PointStruct] = []
    oos_points: list[PointStruct] = []
    vio_idx = 1
    oos_idx = 1

    for idx, row in enumerate(rows, start=1):
        vector = first_vector if idx == 1 else ollama.embed(row.text)
        payload = {
            "text": row.text,
            "kategori": row.category,
            "uu_ref": row.law_ref,
            "source_sheet": row.source_sheet,
            "source_group": row.source_group,
            "chunk_type": row.chunk_type,
            "scenario_hint": row.scenario_hint,
            "sample_input": row.sample_input,
            "keywords": row.keywords,
        }
        if row.source_group == "violation":
            violation_points.append(PointStruct(id=vio_idx, vector=vector, payload=payload))
            vio_idx += 1
        else:
            oos_points.append(PointStruct(id=oos_idx, vector=vector, payload=payload))
            oos_idx += 1

    if violation_points:
        qdrant.upsert(violation_points, settings.qdrant_collection_violation)
    if oos_points:
        qdrant.upsert(oos_points, settings.qdrant_collection_oos)

    return {
        "indexed_total": len(violation_points) + len(oos_points),
        "indexed_violation": len(violation_points),
        "indexed_oos": len(oos_points),
        "message": "Reindex selesai",
    }
